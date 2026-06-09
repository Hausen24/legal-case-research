"""端到端管道冒烟测试：在隔离副本上跑 normalize→analytics→excel，
断言产出存在、Excel 实质列非空、样本量自适应判定正确。"""
import json
import subprocess
import sys

import pytest

GENERAL = "scripts/general"


def run(repo, *args):
    r = subprocess.run([sys.executable, *args], cwd=repo,
                       capture_output=True, text=True)
    assert r.returncode == 0, f"命令失败：{args}\nSTDOUT:{r.stdout}\nSTDERR:{r.stderr}"
    return r


def test_normalize_and_analytics(repo, demo_copy):
    run(repo, f"{GENERAL}/normalize_cases.py", str(demo_copy))
    run(repo, f"{GENERAL}/run_analytics.py", str(demo_copy))

    analytics = json.loads((demo_copy / "06_analytics.json").read_text(encoding="utf-8"))
    assert analytics["样本量"] == 6
    # 小样本 → 定性深挖 + 仅描述性 + 不报告地域分歧（核心方法论保护）
    assert analytics["深度档"]["mode"] == "qualitative_deep"
    assert analytics["定量档"] == "T0_descriptive"
    assert analytics["地域分歧"]["report"] is False
    # 三个争点都应被聚合到
    assert len(analytics["争点出现频次"]) == 3


def test_normalize_derives_norm_fields(repo, demo_copy):
    run(repo, f"{GENERAL}/normalize_cases.py", str(demo_copy))
    cases = json.loads((demo_copy / "05_enriched_cases.json").read_text(encoding="utf-8"))
    for c in cases:
        norm = c.get("_norm", {})
        assert norm.get("法院层级"), f"{c['Gid']} 缺法院层级"
        assert norm.get("裁判年份"), f"{c['Gid']} 缺裁判年份"
        assert norm.get("案由") == "侵害作品信息网络传播权纠纷"


def test_securities_group_litigation_pipeline(repo, demo_sec_copy):
    """集团诉讼（证券示例）管道：normalize→analytics，断言核心机制与样本闸门。"""
    run(repo, "scripts/securities/normalize_secmisrep.py", str(demo_sec_copy))
    run(repo, "scripts/securities/run_analytics_secmisrep.py", str(demo_sec_copy))
    a = json.loads((demo_sec_copy / "06_analytics.json").read_text(encoding="utf-8"))
    assert a["N"] == 7 and a["独立事件数"] == 7
    assert a["深度档"]["mode"] == "qualitative_deep"
    # 上海5/北京2 → 任一组<5 → 不报告地域分歧（核心方法论保护）
    assert a["地域分歧"]["report"] is False


def test_byod_import_builds_pipeline_contract(repo, tmp_path):
    """自带数据：扁平 JSON → 03_raw_cases.json，字段结构与 MCP 一致、可被公共派生解析。"""
    import sys as _sys
    out = tmp_path / "byod"
    run(repo, "scripts/general/import_cases.py", str(out),
        "--json", "examples/byod_sample.json")
    raw = json.loads((out / "03_raw_cases.json").read_text(encoding="utf-8"))
    assert len(raw) == 3
    _sys.path.insert(0, str(repo / "scripts" / "common"))
    import pkulaw_utils as U
    r0 = raw[0]
    assert r0["CaseFlag"] == "（2023）示01民初0001号"
    assert U.flatten_court(r0["LastInstanceCourt"]) == ("北京", "北京互联网法院")
    assert U.flatten_leaf(r0["Category"]) == "侵害作品信息网络传播权纠纷"
    # 普通案例 2 + 公报 1
    codes = sorted(next(iter(r["CaseGrade"])) for r in raw)
    assert codes == ["05", "07", "07"]


def test_excel_substantive_columns_not_empty(repo, demo_copy):
    pytest.importorskip("openpyxl")
    pytest.importorskip("pandas")
    run(repo, f"{GENERAL}/normalize_cases.py", str(demo_copy))
    run(repo, f"{GENERAL}/generate_excel.py", str(demo_copy), "--name", "短视频侵权案件", "--date", "20260607")
    xlsx = demo_copy / "output" / "短视频侵权案件-类案检索清单-20260607.xlsx"
    assert xlsx.exists(), "应按 <案件类别>-类案检索清单-<日期>.xlsx 命名"

    from openpyxl import load_workbook
    wb = load_workbook(xlsx)
    ws = wb["案件清单"]
    header = [c.value for c in ws[1]]
    # 实质列（防"Excel 内容列变空"回归）
    for col in ("案号", "审理法院", "基本案情", "法院裁判要点", "最终裁判结果"):
        assert col in header, f"缺列 {col}"
        idx = header.index(col)
        values = [row[idx].value for row in ws.iter_rows(min_row=2)]
        assert all(v not in (None, "") for v in values), f"列“{col}”存在空单元格"
    assert "权威案例附录" in wb.sheetnames
