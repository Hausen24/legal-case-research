#!/usr/bin/env python3
"""
generate_excel_secmisrep.py —— 证券虚假陈述核心判决清单（长表三 sheet）
用法：python3 generate_excel_secmisrep.py <research_dir>
读 <research_dir>/05_enriched_cases.json（含 问题观点 dict），按 excel-schema.md 产出
output/核心判决清单.xlsx：Sheet1 判决要点（长表）/ Sheet2 争点编码（tidy）/ Sheet3 案例索引。
平行/典型来自 04_screened_cases.json（_track∈{parallel,typical}），缺则略。
"""
import sys, os, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY="1F3864"; BAND="EEF1F6"; LINE="C9C9C9"; EA="宋体"; EAH="微软雅黑"
thin=Side(style="thin",color=LINE); BORD=Border(left=thin,right=thin,top=thin,bottom=thin)
def F(sz=10.5,b=False,color="222222",name=EA): return Font(name=name,size=sz,bold=b,color=color)
WT=Alignment(wrap_text=True,vertical="top"); WC=Alignment(wrap_text=True,vertical="center",horizontal="center")
HD=Alignment(wrap_text=True,vertical="center",horizontal="center")

def header(ws, cols):
    for ci,(name,w) in enumerate(cols,1):
        c=ws.cell(row=1,column=ci,value=name); c.font=F(10.5,True,"FFFFFF",EAH)
        c.fill=PatternFill("solid",fgColor=NAVY); c.alignment=HD; c.border=BORD
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[1].height=32

def output_name(argv):
    """据 --name/--date 生成清单文件名：<案件类别>-类案检索清单-<YYYYMMDD>.xlsx。
    未给 --name 时回退旧名"核心判决清单.xlsx"（向后兼容）。"""
    import datetime
    name,date="",""
    i=2
    while i<len(argv):
        if argv[i]=="--name" and i+1<len(argv): name=argv[i+1].strip(); i+=2
        elif argv[i]=="--date" and i+1<len(argv): date=argv[i+1].strip(); i+=2
        else: i+=1
    if not name: return "核心判决清单.xlsx"
    if not date: date=datetime.date.today().strftime("%Y%m%d")
    return f"{name}-类案检索清单-{date}.xlsx"


def main():
    if len(sys.argv)<2:
        sys.exit('用法：python3 scripts/securities/generate_excel_secmisrep.py <research_dir> '
                 '[--name "<案件类别>"] [--date YYYYMMDD]')
    rd=sys.argv[1]
    cases=json.load(open(os.path.join(rd,"05_enriched_cases.json"),encoding="utf-8"))
    out_dir=os.path.join(rd,"output"); os.makedirs(out_dir,exist_ok=True)
    wb=Workbook()

    # ---- Sheet1 判决要点（长表）----
    ws=wb.active; ws.title="判决要点"
    cols=[("序号",6),("案号",18),("案件名称",26),("涉案上市公司",16),("核心判决类型",12),
          ("审理法院",14),("审级",9),("裁判日期",12),("裁判结果（总）",11),
          ("基本案情（时间·地点·主体—起因·经过·结果）",52),
          ("争议焦点",16),("各方抗辩主张及理由",40),("裁判观点及理由",46),("案件链接",14)]
    header(ws, cols)
    CASE_COLS=[1,2,3,4,5,6,7,8,9,10,14]
    r=2; band=False
    for c in cases:
        qv=c.get("问题观点") or {}
        items=list(qv.items()) or [("—",{"裁判观点":"（该案争点编码略）"})]
        start=r
        link=c.get("北大法宝链接") or c.get("Url") or ""
        meta={1:c.get("序号"),2:c.get("案号"),3:c.get("案件名称"),4:c.get("涉案上市公司"),
              5:c.get("核心判决类型"),6:c.get("审理法院"),7:c.get("审级"),
              8:c.get("裁判日期"),9:c.get("裁判结果分类"),10:c.get("基本案情") or "",14:"查看原文"}
        fill=PatternFill("solid",fgColor=BAND) if band else None
        for (issue,info) in items:
            for ci in CASE_COLS:
                cell=ws.cell(row=r,column=ci,value=(meta[ci] if r==start else None))
                cell.border=BORD; cell.font=F(); cell.alignment=(WT if ci in (3,10) else WC)
                if fill: cell.fill=fill
            if link:
                lc=ws.cell(row=start,column=14); lc.hyperlink=link; lc.font=F(10.5,False,"1F4E79")
            tag=f"〔{info.get('倾向标签')}〕\n" if info.get("倾向标签") and info.get("倾向标签")!="—" else ""
            vals={11:info.get("争议焦点") or issue,
                  12:info.get("各方抗辩") or "",
                  13:(tag+(info.get("裁判观点") or "")).strip()}
            for ci,v in vals.items():
                cell=ws.cell(row=r,column=ci,value=v); cell.border=BORD; cell.font=F(); cell.alignment=WT
                if fill: cell.fill=fill
            ws.row_dimensions[r].height=max(46,16*(1+len(str(vals[13]))//22)); r+=1
        if r-1>start:
            for ci in CASE_COLS:
                ws.merge_cells(start_row=start,start_column=ci,end_row=r-1,end_column=ci)
        band=not band
    ws.freeze_panes="C2"; ws.auto_filter.ref=f"A1:{get_column_letter(len(cols))}1"

    # ---- Sheet2 争点编码（tidy）----
    ws2=wb.create_sheet("争点编码")
    h2=[("案号",18),("审理法院",14),("法院地",8),("裁判年份",10),("争点类别",18),
        ("裁判倾向标签",24),("裁判结果分类",11),("是否典型",16),("相关度",8)]
    header(ws2,h2); rr=2
    for c in cases:
        for issue,info in (c.get("问题观点") or {}).items():
            vals=[c.get("案号"),c.get("审理法院"),c.get("法院地"),c.get("裁判年份"),issue,
                  info.get("倾向标签") or "—",c.get("裁判结果分类"),c.get("是否典型案例") or "",c.get("相关度") or ""]
            for ci,v in enumerate(vals,1):
                cell=ws2.cell(row=rr,column=ci,value=v); cell.border=BORD; cell.font=F(10)
                cell.alignment=(WT if ci in (5,6) else WC)
                if rr%2==0: cell.fill=PatternFill("solid",fgColor=BAND)
            ws2.row_dimensions[rr].height=26; rr+=1
    ws2.freeze_panes="A2"; ws2.auto_filter.ref=f"A1:{get_column_letter(len(h2))}1"

    # ---- Sheet3 案例索引（平行+典型，来自 04）----
    ws3=wb.create_sheet("案例索引")
    h3=[("序号",6),("类别",10),("案件名称",30),("案号",20),("审理法院/名录",24),("链接",12)]
    header(ws3,h3); rr=2; idx=1
    p04=os.path.join(rd,"04_screened_cases.json")
    aux=[x for x in (json.load(open(p04,encoding="utf-8")) if os.path.exists(p04) else [])
         if x.get("_track") in ("parallel","typical")]
    for x in aux:
        cat="平行判决" if x.get("_track")=="parallel" else "典型/参考"
        vals=[idx,cat,x.get("Title") or x.get("案件名称",""),x.get("CaseFlag") or x.get("案号",""),
              x.get("入选名录") or x.get("所属核心案",""),"查看"]
        for ci,v in enumerate(vals,1):
            cell=ws3.cell(row=rr,column=ci,value=v); cell.border=BORD; cell.font=F(10)
            cell.alignment=(WT if ci==3 else WC)
            if rr%2==0: cell.fill=PatternFill("solid",fgColor=BAND)
        url=x.get("Url") or x.get("链接")
        if url:
            lc=ws3.cell(row=rr,column=6); lc.hyperlink=url; lc.font=F(10,False,"1F4E79")
        ws3.row_dimensions[rr].height=24; rr+=1; idx+=1
    ws3.freeze_panes="A2"; ws3.auto_filter.ref=f"A1:{get_column_letter(len(h3))}1"

    path=os.path.join(out_dir,output_name(sys.argv)); wb.save(path)
    print("Excel 已生成：",path,"| 判决要点行",ws.max_row,"争点编码行",ws2.max_row)

if __name__=="__main__":
    main()
