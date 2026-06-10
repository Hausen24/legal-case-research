"""
generate_excel.py —— 生成案件清单 Excel（通用类案研究版）
─────────────────────────────────────────────
用法：python3 scripts/general/generate_excel.py <research_dir>
输入：
  <research_dir>/05_enriched_cases.json   分析池（普通案例，已编码+规范化）
  <research_dir>/03_raw_cases.json        检索原始累加（字段回退 + 权威案例附录来源）
  <research_dir>/04_screened_cases.json   （可选）若其中含非07案例，则并入权威案例附录
输出：<research_dir>/output/案件清单.xlsx
依赖：pip3 install pandas openpyxl
公共字段处理函数集中在 scripts/common/pkulaw_utils.py（含 clean_url 抠裸链接）。
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "common"))
from pkulaw_utils import (  # noqa: E402
    load, clean_url, derive_court_level, flatten_court, flatten_leaf,
    derive_year, build_raw_index, field,
)

FOCUS_ICONS = "①②③④⑤⑥⑦⑧⑨⑩"


def resolve_norm(c: dict, raw_idx: dict) -> dict:
    """取 _norm；任一派生列缺失时，用原始字段就地重算兜底（防止 normalize 漏跑导致整列空白）。"""
    norm = dict(c.get("_norm") or {})
    prov_fb, court_fb = flatten_court(field(c, raw_idx, "LastInstanceCourt", {}))
    court = norm.get("法院全称") or court_fb
    norm["法院全称"] = court
    norm["法院层级"] = norm.get("法院层级") or derive_court_level(court)
    norm["地域"] = norm.get("地域") or prov_fb
    norm["裁判年份"] = norm.get("裁判年份") or derive_year(field(c, raw_idx, "LastInstanceDate", ""))
    norm["案由"] = norm.get("案由") or flatten_leaf(field(c, raw_idx, "Category", {}))
    norm["案件等级"] = norm.get("案件等级") or (flatten_leaf(field(c, raw_idx, "CaseGrade", {})) or "普通案例")
    return norm


def collect_focuses(cases):
    """收集本次研究出现的所有争议焦点名（保持稳定顺序）"""
    seen = []
    for c in cases:
        for f in (c.get("焦点立场") or {}).keys():
            if f not in seen:
                seen.append(f)
    return seen


def build_main_rows(cases, focuses, raw_idx):
    rows = []
    for i, c in enumerate(cases, 1):
        norm = resolve_norm(c, raw_idx)
        row = {
            "序号": i,
            "案件名称": field(c, raw_idx, "Title"),
            "案号": field(c, raw_idx, "CaseFlag"),
            "审理法院": norm.get("法院全称", ""),
            "法院层级": norm.get("法院层级", ""),
            "地域": norm.get("地域", ""),
            "审级": field(c, raw_idx, "CaseClassName"),
            "案件等级": norm.get("案件等级", ""),
            "案由": norm.get("案由", ""),
            "裁判日期": field(c, raw_idx, "LastInstanceDate"),
            "裁判年份": norm.get("裁判年份", ""),
            # 浓缩摘要列（Claude Code 编码产出；为空则回退原始字段）
            "基本案情": c.get("基本案情摘要") or field(c, raw_idx, "Ascertain"),
            "原告诉求": field(c, raw_idx, "PlaintiffClaims"),
            "辩方抗辩要点": c.get("抗辩要点摘要") or field(c, raw_idx, "DefenseViewpoint"),
            "争议焦点": field(c, raw_idx, "ControversialFocus"),
            "法院裁判要点": c.get("裁判要点摘要") or field(c, raw_idx, "Identified"),
            "裁判依据": field(c, raw_idx, "RefereeBasis"),
            "最终裁判结果": field(c, raw_idx, "RefereeResult"),
            "裁判结果分类": c.get("裁判结果分类", ""),
            "判赔金额(元)": c.get("判赔金额", ""),
            "维权开支(元)": c.get("维权开支", ""),
            "相关度": c.get("相关度", ""),
        }
        positions = c.get("焦点立场") or {}
        for j, f in enumerate(focuses):
            icon = FOCUS_ICONS[j] if j < len(FOCUS_ICONS) else f"{j+1}."
            info = positions.get(f, {})
            row[f"{icon}{f}-立场"] = info.get("立场", "")
            row[f"{icon}{f}-理由"] = info.get("理由", "")
        row["北大法宝链接"] = clean_url(field(c, raw_idx, "Url"))
        rows.append(row)
    return rows


def build_appendix_rows(raw_idx, screened):
    """权威案例附录：所有检索到的非07案例（经典/评析/指导/公报等），按 Gid 去重。
    以 03 去重后的全集为基底，并并入 04 中可能多出的非07记录。"""
    candidates = dict(raw_idx)
    for c in screened:
        gid = c.get("Gid")
        if gid and gid not in candidates:
            candidates[gid] = c

    rows, i = [], 1
    for rec in candidates.values():
        grade = rec.get("CaseGrade")
        if not isinstance(grade, dict) or not grade:
            continue
        if "07" in grade:
            continue
        rows.append({
            "序号": i,
            "案件名称": rec.get("Title", ""),
            "案号": rec.get("CaseFlag", ""),
            "案件等级": flatten_leaf(grade),
            "北大法宝链接": clean_url(rec.get("Url", "")),
        })
        i += 1
    return rows


def build_divergence_rows(rd):
    """从 06_analytics.json 的 分歧地图 生成「裁判分歧清单」行（无 06 或无分歧则空）。"""
    p = rd / "06_analytics.json"
    if not p.exists():
        return []
    import json as _json
    div = (_json.loads(p.read_text(encoding="utf-8")) or {}).get("分歧地图") or {}
    rows = []
    for i, (issue, info) in enumerate(div.items(), 1):
        stances = sorted((info.get("立场分布") or {}).items(), key=lambda x: -x[1])
        if len(stances) < 2:
            continue
        (sa, na), (sb, nb) = stances[0], stances[1]
        reps = info.get("代表案") or {}
        rows.append({
            "序号": i,
            "争议焦点": issue,
            "立场A(件数)": f"{sa}（{na}）",
            "立场A代表案号": "；".join(reps.get(sa, [])),
            "立场B(件数)": f"{sb}（{nb}）",
            "立场B代表案号": "；".join(reps.get(sb, [])),
            "其他立场": "；".join(f"{s}（{n}）" for s, n in stances[2:]),
            "样本提示": info.get("措辞", ""),
        })
    return rows


def write_excel(main_rows, appendix_rows, out_path, divergence_rows=None):
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    out_path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(main_rows)

    long_cols = {"基本案情", "原告诉求", "辩方抗辩要点", "争议焦点",
                 "法院裁判要点", "裁判依据", "最终裁判结果"}

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="案件清单")
        ws = writer.sheets["案件清单"]

        hdr_fill = PatternFill("solid", fgColor="1F3864")
        hdr_font = Font(color="FFFFFF", bold=True, size=11)
        for i, cell in enumerate(ws[1]):
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            col_name = df.columns[i]
            width = 42 if col_name in long_cols else (6 if col_name == "序号" else 16)
            ws.column_dimensions[get_column_letter(i + 1)].width = width
        ws.row_dimensions[1].height = 30

        fill_even = PatternFill("solid", fgColor="EBF2FF")
        for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            if r_idx % 2 == 0:
                for cell in row:
                    cell.fill = fill_even
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[r_idx].height = 90
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        if appendix_rows:
            df2 = pd.DataFrame(appendix_rows)
            df2.to_excel(writer, index=False, sheet_name="权威案例附录")
            ws2 = writer.sheets["权威案例附录"]
            for i, cell in enumerate(ws2[1]):
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws2.column_dimensions[get_column_letter(i + 1)].width = [6, 45, 24, 22, 50][i] if i < 5 else 18
            for r_idx, row in enumerate(ws2.iter_rows(min_row=2), start=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws2.freeze_panes = "A2"

        if divergence_rows:
            df3 = pd.DataFrame(divergence_rows)
            df3.to_excel(writer, index=False, sheet_name="裁判分歧清单")
            ws3 = writer.sheets["裁判分歧清单"]
            for i, cell in enumerate(ws3[1]):
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws3.column_dimensions[get_column_letter(i + 1)].width = \
                    [6, 26, 16, 26, 16, 26, 18, 40][i] if i < 8 else 18
            for r_idx, row in enumerate(ws3.iter_rows(min_row=2), start=2):
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                ws3.row_dimensions[r_idx].height = 48
            ws3.freeze_panes = "A2"

    print(f"Excel 已生成：{out_path}")
    print(f"  案件清单 {len(main_rows)} 条 | 权威案例附录 {len(appendix_rows)} 条 | "
          f"裁判分歧清单 {len(divergence_rows or [])} 条")


def output_name(argv):
    """据 --name/--date 生成清单文件名：<案件类别>-类案检索清单-<YYYYMMDD>.xlsx。
    未给 --name 时回退旧名"案件清单.xlsx"（向后兼容）。"""
    import datetime
    name, date = "", ""
    i = 2
    while i < len(argv):
        if argv[i] == "--name" and i + 1 < len(argv):
            name = argv[i + 1].strip(); i += 2
        elif argv[i] == "--date" and i + 1 < len(argv):
            date = argv[i + 1].strip(); i += 2
        else:
            i += 1
    if not name:
        return "案件清单.xlsx"
    if not date:
        date = datetime.date.today().strftime("%Y%m%d")
    return f"{name}-类案检索清单-{date}.xlsx"


def main():
    if len(sys.argv) < 2:
        sys.exit('用法：python3 scripts/general/generate_excel.py <research_dir> '
                 '[--name "<案件类别>"] [--date YYYYMMDD]')
    rd = Path(sys.argv[1])
    cases = load(rd / "05_enriched_cases.json")
    if not cases:
        sys.exit(f"未找到或为空：{rd/'05_enriched_cases.json'}")
    screened = load(rd / "04_screened_cases.json")
    raw_idx = build_raw_index(rd)

    focuses = collect_focuses(cases)
    main_rows = build_main_rows(cases, focuses, raw_idx)
    appendix_rows = build_appendix_rows(raw_idx, screened)
    write_excel(main_rows, appendix_rows, rd / "output" / output_name(sys.argv),
                divergence_rows=build_divergence_rows(rd))


if __name__ == "__main__":
    main()
